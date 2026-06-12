# =============================================
# app/services/excel_report.py
# 역할: 하자점검 결과 보고서를 '엑셀 양식'(app/templates/defect_report_template.xlsx)에
#       그대로 채워서 생성. 시트1=점검개요+하자 상세내역 표, 시트2=하자 사진 첨부(이미지 삽입).
#       기존 마크다운 LLM 보고서(llm_report.py)와 별개로, 실제 제출용 양식 산출물.
#
#  매핑(우리 20종 taxonomy → 양식):
#    - 분류코드 컬럼 = 양식의 12분류(A.도장·도배 ~ L.기타) 중 하나로 매핑
#    - 등급 컬럼 = 심각도 HIGH→C(중대) / MED→B(보통) / LOW→A(경미)
# =============================================
from __future__ import annotations

import base64
import io
import os
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

_TEMPLATE = os.path.join(os.path.dirname(__file__), "..", "templates", "defect_report_template.xlsx")
_SHEET_MAIN = "하자점검 결과보고서"
_SHEET_PHOTO = "하자 사진 첨부"

# 우리 category_code → 양식 분류코드(A~L)
_CODE_TO_CAT = {
    "A-01": "I", "A-02": "I", "A-03": "I", "A-04": "D", "A-05": "I",
    "B-01": "D", "B-02": "L", "B-03": "J", "B-04": "J", "B-05": "D",
    "C-01": "A", "C-02": "A", "C-03": "A", "C-04": "A", "C-05": "C", "C-06": "A",
    "D-01": "G", "D-02": "B", "D-03": "B", "D-04": "B",
    "E-01": "D", "E-02": "D",
}
_CAT_NAME = {
    "A": "도장·도배", "B": "타일·석재", "C": "목공·수장", "D": "창호",
    "E": "금속·잡철", "F": "전기·조명", "G": "기계·설비", "H": "위생·급배수",
    "I": "구조·균열", "J": "방수·누수", "K": "조경·외구", "L": "기타",
}
# 위치/부위 추정(룸 정보가 없으므로 하자 종류 기반 coarse 힌트 — 점검자가 보정)
_CODE_TO_LOC = {
    "A-01": "벽·천장", "A-02": "구조부", "A-03": "벽·천장", "A-04": "창호부", "A-05": "구조부",
    "B-01": "창호부", "B-02": "벽체", "B-03": "코킹부(창호·욕실)", "B-04": "방수부위", "B-05": "창호부",
    "C-01": "벽(도배)", "C-02": "벽(도배)", "C-03": "벽·천장", "C-04": "벽·천장", "C-05": "걸레받이", "C-06": "벽·천장 마감",
    "D-01": "바닥(난방)", "D-02": "바닥", "D-03": "바닥", "D-04": "바닥(줄눈)",
    "E-01": "창유리", "E-02": "창틀·문틀",
}
_SEV_TO_GRADE = {"HIGH": "C", "MED": "B", "LOW": "A"}
_GRADE_ACTION = {"C": "구조·안전 점검 후 즉시 시정", "B": "해당 부위 재시공", "A": "단순 마감 보수"}
_GRADE_DEADLINE = {"C": "즉시(7일 이내)", "B": "14일 이내", "A": "30일 이내"}

# 시트1 하자 상세내역 표: 데이터 시작행 19, 10행(19~28)
_DETAIL_START_ROW = 19
_DETAIL_MAX_ROWS = 10


def _grade(sev: Optional[str]) -> str:
    return _SEV_TO_GRADE.get((sev or "LOW").upper(), "A")


def _cat(code: Optional[str]) -> str:
    return _CODE_TO_CAT.get(code or "", "L")


def _decode_image(image_crop) -> Optional[PILImage.Image]:
    """base64(또는 data URL) → PIL 이미지. 실패 시 None."""
    if not image_crop:
        return None
    try:
        s = image_crop
        if isinstance(s, str) and s.startswith("data:"):
            s = s.split(",", 1)[1]
        raw = base64.b64decode(s) if isinstance(s, str) else s
        return PILImage.open(io.BytesIO(raw)).convert("RGB")
    except Exception:
        return None


def _xl_image(pil: PILImage.Image, target_w: int = 230) -> XLImage:
    """PIL → openpyxl 삽입 이미지(가로 target_w px 로 축소)."""
    w, h = pil.size
    if w > target_w:
        pil = pil.resize((target_w, max(1, int(h * target_w / w))))
    buf = io.BytesIO()
    pil.save(buf, format="PNG")
    buf.seek(0)
    return XLImage(buf)


def build_excel_report(
    defects: List[dict],
    *,
    site_name: str = "",
    unit: str = "",
    inspector: str = "",
    inspect_area: str = "",
) -> bytes:
    """검출 하자 리스트 → 양식 채운 xlsx 바이트.

    defects: [{category_code, defect_type, severity, confidence, image_crop(base64), ...}]
             심각도 내림차순 권장. image_crop 있으면 시트2에 사진 첨부.
    """
    wb = openpyxl.load_workbook(_TEMPLATE)
    ws = wb[_SHEET_MAIN]

    # ── 점검 개요 ──
    today = datetime.now()
    if site_name:
        ws["C5"] = site_name
    if unit:
        ws["C6"] = unit
    ws["C7"] = today.strftime("%Y년 %m월 %d일")
    if inspector:
        ws["C8"] = inspector
    if inspect_area:
        ws["G6"] = f"{inspect_area} ㎡"

    # ── 하자 상세내역(시트1) ──
    rows = defects[:_DETAIL_MAX_ROWS]
    for i, d in enumerate(rows):
        r = _DETAIL_START_ROW + i
        code = d.get("category_code") or (d.get("defect_info") or {}).get("category_code")
        name = d.get("defect_type") or (d.get("defect_info") or {}).get("defect_type") or ""
        sev = d.get("severity") or (d.get("defect_info") or {}).get("severity")
        g = _grade(sev)
        ws.cell(r, 2, _cat(code))                         # B 분류코드(A~L)
        ws.cell(r, 3, _CODE_TO_LOC.get(code, ""))         # C 위치/부위
        ws.cell(r, 4, f"[{code}] {name}")                 # D 하자내용(merged D:E)
        ws.cell(r, 6, g)                                  # F 등급(A/B/C)
        ws.cell(r, 7, _GRADE_ACTION[g])                   # G 조치방법
        ws.cell(r, 8, _GRADE_DEADLINE[g])                 # H 처리기한

    if len(defects) > _DETAIL_MAX_ROWS:
        # 양식 10행 초과분은 종합의견에 표기(점검자 인지용)
        ws["A33"] = (f"※ 총 검출 하자 {len(defects)}건 중 상위 {_DETAIL_MAX_ROWS}건을 "
                     f"상세표에 기재. 전체 목록·사진은 첨부 시트 참조.")

    # ── 하자 사진 첨부(시트2) — 동적 2열 그리드로 재구성 ──
    _fill_photos(wb[_SHEET_PHOTO], [d for d in defects if _decode_image(d.get("image_crop"))])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fill_photos(ws, photo_defects: List[dict]) -> None:
    """시트2를 2열 사진 그리드로 채운다(템플릿 샘플 블록 영역 재사용/확장).
    각 블록: 이미지 + '사진 N / 위치 / 분류코드 / 설명'."""
    # 헤더(행 1~2) 아래 병합은 모두 해제 — 병합 셀에 쓰면 read-only 에러.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 4:
            ws.unmerge_cells(str(rng))
    # 기존 샘플 안내 텍스트(행 4 이하) 정리
    for r in range(4, ws.max_row + 1):
        for c in range(1, 8):
            ws.cell(r, c).value = None

    band_h = 9          # 한 밴드(사진+라벨) 높이(행)
    img_rows = 6        # 이미지가 차지하는 행 수
    left_col, right_col = 1, 5   # A, E
    row = 4
    for i in range(0, len(photo_defects), 2):
        for slot, col in ((i, left_col), (i + 1, right_col)):
            if slot >= len(photo_defects):
                break
            d = photo_defects[slot]
            pil = _decode_image(d.get("image_crop"))
            code = d.get("category_code") or (d.get("defect_info") or {}).get("category_code") or ""
            name = d.get("defect_type") or (d.get("defect_info") or {}).get("defect_type") or ""
            col_l = get_column_letter(col)
            # 이미지
            if pil is not None:
                try:
                    ws.add_image(_xl_image(pil), f"{col_l}{row}")
                except Exception:
                    pass
            # 라벨(이미지 아래)
            lr = row + img_rows
            ws.cell(lr, col, f"사진 {slot + 1}")
            ws.cell(lr, col + 1, f"위치: {_CODE_TO_LOC.get(code, '-')}")
            ws.cell(lr + 1, col, "분류코드")
            ws.cell(lr + 1, col + 1, f"{_cat(code)}. {_CAT_NAME.get(_cat(code), '')} ([{code}])")
            ws.cell(lr + 2, col, "설명")
            ws.cell(lr + 2, col + 1, name)
        row += band_h
    # 이미지가 들어가는 행 높이 확보
    for r in range(4, row + 1):
        ws.row_dimensions[r].height = 18
